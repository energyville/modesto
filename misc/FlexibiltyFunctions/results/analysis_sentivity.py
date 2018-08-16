import os
import pickle
import pandas as pd
import modesto.utils as ut
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pandas import ExcelWriter
import numpy as np
import math
import openpyxl


class SensitivityAnalysis:

    def __init__(self, name, source_dir, target_dir):
        self.name = name
        self.source_dir = source_dir
        self.target_dir = target_dir
        self.makedir(target_dir)

        self.data = self.load_data()
        self.sens_cases = self.data['selected_sensitivity_cases']
        self.model_cases = self.data['selected_model_cases']
        self.neigh_cases = self.data['selected_street_cases'] + self.data['selected_district_cases']
        self.sens_values = self.data['sensitivity_values']
        self.solutions = ['energy_use_difference', 'energy_use_ref', 'energy_use_flex',
                          'energy_cost_ref', 'energy_cost_flex', 'energy_cost_difference',
                          'upward_energy', 'downward_energy']

        self.results = self.load_results()
        self.pit = self.find_price_increase_time(self.data)

        self.output = None

    def load_obj(self, file_loc):
        with open(file_loc, 'rb') as fp:
            return pickle.load(fp)

    def load_data(self):
        return self.load_obj(os.path.join(self.source_dir, self.name + '_data.pkl'))

    def load_results(self):
        results = {}
        for sens_case in self.sens_cases:
            results[sens_case] = self.load_obj(os.path.join(self.source_dir, self.name + '_' + sens_case + '.pkl'))
        return results

    def makedir(self, name):
        if not os.path.exists(name):
            os.makedirs(name)

    def make_dataframe(self, sens_case, index1, columns):
        index2 = self.sens_values[sens_case]
        index1_lists = [[key]*len(index2) for key in index1]
        index1_list = [item for sublist in index1_lists for item in sublist]
        index2_lists = [index2*len(index1)]
        index2_list = [item for sublist in index2_lists for item in sublist]
        tuples = list(zip(*[index1_list, index2_list]))
        index = pd.MultiIndex.from_tuples(tuples, names=['Result type', 'Sensitivity values'])
        return pd.DataFrame(index=index, columns=columns)

    def save_xls(self, dict_dfs, file_name):
        xls_path = os.path.join(self.target_dir, file_name)
        if not os.path.isfile(xls_path):
            wb = openpyxl.Workbook()
            wb.save(xls_path)
        book = openpyxl.load_workbook(xls_path)

        for sheet in book.sheetnames:
            if 'Sheet' in sheet:
                book.remove(book[sheet])

        writer = ExcelWriter(xls_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for name, df in dict_dfs.items():
            df.to_excel(writer, name)

        writer.save()

    def save_output(self):
        if self.output is None:
            raise Exception('No output has been collected yet')
        self.save_xls(self.output, sim_name + '.xlsx')

    def savefig(self, figure, figname):
        figure.savefig(os.path.join(self.target_dir, self.name, figname + '.svg'))

    def find_result(self, sens_case=None, sens_val=None, neigh_case=None, model_case=None, flex_case=None, var_name=None):
        if sens_case is None:
            return self.results
        elif neigh_case is None:
            return self.results[sens_case]
        elif model_case is None:
            return self.results[sens_case][neigh_case]
        elif sens_val is None:
            return self.results[sens_case][neigh_case][sens_val]
        elif flex_case is None:
            return self.results[sens_case][neigh_case][sens_val][model_case]
        elif var_name is None:
            return self.results[sens_case][neigh_case][sens_val][model_case][flex_case]
        else:
            return self.results[sens_case][neigh_case][sens_val][model_case][flex_case][var_name]

    def energy_use_kwh(self, heat_injection, time_step):  # TODO Resample price
        return sum(heat_injection) * time_step /1000/3600

    def difference(self, use1, use2):
        return use1 - use2

    def resample(self, data, new_time_step, last_point):
        resampled = data.resample(str(new_time_step) + 'S').pad()
        resampled = resampled.ix[~(resampled.index > last_point)]

        return resampled

    def energy_cost(self, heat_injection, price, time_step):
        price_profile = price.multiply(heat_injection) * time_step / 1000 / 3600
        return sum(price_profile[np.isfinite(price_profile)])

    def find_price_increase_time(self, data):
        price = data['price_profiles']['step']

        position = price.index[
            next(x[0] for x in enumerate(price) if x[1] == 2)]

        return position

    def find_time_step(self, model_case):
        return self.data['model_cases'][model_case]['time_step']

    def find_last_point(self):
        return pd.Timestamp(self.data['start_time'] + pd.Timedelta(seconds=self.data['horizon']))

    def upward_downward_power_kwh(self, heat_injection_flex, heat_injection_ref, pit, time_step):
        response = self.difference(heat_injection_flex,
                                   heat_injection_ref)

        return sum(response.ix[response.index < pit]) * time_step / 1000 / 3600, \
               -sum(response.ix[response.index >= pit]) * time_step / 1000 / 3600

    def collect_results(self, sens_cases=None, neigh_cases=None, model_cases=None):
        if sens_cases is None:
            sens_cases = self.sens_cases
        if neigh_cases is None:
            neigh_cases = self.neigh_cases
        if model_cases is None:
            model_cases = self.model_cases

        output = {}

        for sens in sens_cases:
            output[sens] = self.make_dataframe(sens, self.solutions, model_cases)

        for sens in sens_cases:
            for sens_val in self.sens_values[sens]:
                for neigh in neigh_cases:
                    for model in model_cases:

                        def change_element(df, new_val, sol_type):
                            df.loc[(sol_type, sens_val), model] = new_val
                            return df

                        def get_element(df, sol_type):
                            return df.loc[(sol_type, sens_val), model]

                        case_results_flex = self.find_result(sens_case=sens,
                                                             neigh_case=neigh,
                                                             model_case=model,
                                                             flex_case='Flexibility',
                                                             sens_val=sens_val)
                        case_results_ref = self.find_result(sens_case=sens,
                                                            neigh_case=neigh,
                                                            model_case=model,
                                                            flex_case='Reference',
                                                            sens_val=sens_val)

                        if case_results_flex is not None and case_results_ref is not None:
                            sens_output = output[sens]
                            time_step = self.find_time_step(model)
                            price = self.resample(self.data['price_profiles']['step'], time_step, self.find_last_point())
                            sens_output = change_element(
                                sens_output,
                                self.energy_use_kwh(case_results_ref['heat_injection'], time_step), 'energy_use_ref')
                            sens_output = change_element(
                                sens_output,
                                self.energy_use_kwh(case_results_flex['heat_injection'], time_step), 'energy_use_flex')
                            sens_output = change_element(
                                sens_output,
                                self.difference(get_element(sens_output, 'energy_use_flex'),
                                                get_element(sens_output, 'energy_use_ref')),
                                'energy_use_difference')
                            sens_output = change_element(
                                sens_output,
                                self.energy_cost(case_results_ref['heat_injection'], price, time_step),
                                'energy_cost_ref')
                            sens_output = change_element(
                                sens_output,
                                self.energy_cost(case_results_flex['heat_injection'], price, time_step),
                                'energy_cost_flex')
                            sens_output = change_element(
                                sens_output,
                                self.difference(get_element(sens_output, 'energy_cost_ref'),
                                                get_element(sens_output, 'energy_cost_flex')),
                                'energy_cost_difference')

                            upward_energy, downward_energy = self.upward_downward_power_kwh(
                                case_results_flex['heat_injection'], case_results_ref['heat_injection'], self.pit, time_step)
                            sens_output = change_element(sens_output, upward_energy,'upward_energy')
                            sens_output = change_element(sens_output, downward_energy, 'downward_energy')

                        else:
                            print 'WARNING: There is no data for case {}: {} - {} - {}'.format(sens, sens_val, neigh, model)
        self.output = output
        return output


if __name__ == '__main__':

    sim_name = 'Sensitivity_analysis_2806'
    source_dir = os.path.dirname(os.path.realpath(__file__))
    target_dir = os.path.join(source_dir, sim_name)

    analysis = SensitivityAnalysis(sim_name, source_dir, target_dir)
    print analysis.collect_results()
    analysis.save_output()



    #
    # def save_obj(obj, name):
    #     with open(name + '.pkl', 'wb') as f:
    #         pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)
    #
    # import pandas as pd
    # pos = 2.75 / 7
    #
    # data = {'n_buildings': 10,
    #         'n_streets': 3,
    #         'horizon': 24 * 7 * 3600,
    #         'start_time': pd.Timestamp('20140101'),
    #         'selected_flex_cases': ['Reference', 'Flexibility'],
    #         'selected_model_cases': ['Buildings', 'Network', 'Combined - LP'],
    #         'selected_street_cases': [],
    #         'selected_district_cases': ['Genk'],
    #         'selected_sensitivity_cases': ['network_size', 'pipe_lengths', 'pipe_diameters',
    #                                        'heat_demand', 'supply_temp_level', 'supply_temp_reach',
    #                                        'substation_temp_difference'],
    #         'sensitivity_values': {'network_size': [0.001, 0.01, 0.1, 1],
    #                                'pipe_lengths': [0.8, 0.9, 1, 1.1, 1.2],
    #                                'pipe_diameters': [-1, 0, 1],
    #                                'heat_demand': [0.8, 0.9, 0.95, 1, 1.05, 1.1, 1.2],
    #                                'supply_temp_level': [i + 273.15 for i in [50, 60, 70, 80, 90]],
    #                                'supply_temp_reach': [5, 7.5, 10, 12.5, 15, 17.5, 20],
    #                                'substation_temp_difference': [15, 20, 25, 30, 35, 40]},
    #         'dist_pipe_length': 150,
    #         'street_pipe_length': 30,
    #         'service_pipe_length': 30,
    #         'old_building': 'SFH_D_3_2zone_TAB',
    #         'mixed_building': 'SFH_D_3_2zone_REF1',
    #         'new_building': 'SFH_D_5_ins_TAB',
    #         'streets': ['Old street', 'Mixed street', 'New street'],
    #         'districts': ['Series district', 'Parallel district', 'Genk'],
    #         'models': ['Buildings - ideal network', 'Buildings', 'Network', 'Combined - LP'],
    #         'max_heat': {'SFH_D_5_ins_TAB': 8900,
    #                      'SFH_D_3_2zone_REF1': 8659,
    #                      'SFH_D_3_2zone_TAB': 20000},
    #         'time_steps': {'StSt': 900,
    #                        'Dynamic': 300},
    #         'pipe_models': {'NoPipes': 'SimplePipe',
    #                         'StSt': 'ExtensivePipe',
    #                         'Dynamic': 'NodeMethod'},
    #         'building_models': {'RC': 'RCmodel',
    #                             'Fixed': 'BuildingFixed'}
    #
    #         }
    #
    # data['time_index'] = pd.date_range(start=data['start_time'], periods=int(data['horizon'] / 3600) + 1, freq='H')
    #
    # data['building_types'] = {'Mixed street': [data['new_building'], data['old_building']] * int(data['n_buildings'] / 2),
    #                           'Old street': [data['old_building']] * data['n_buildings'],
    #                           'New street': [data['new_building']] * data['n_buildings'],
    #                           'Series district': [data['old_building'], data['mixed_building'], data['new_building']],
    #                           'Parallel district': [data['old_building'], data['mixed_building'], data['new_building']],
    #                           'Genk': [data['old_building']] * 9}
    #
    # data['price_profiles'] = {'constant': pd.Series(1, index=data['time_index']),
    #                           'step': pd.Series([1] * int(len(data['time_index']) * pos) +
    #                                      [2] * (len(data['time_index']) - int(len(data['time_index']) * pos)),
    #                                      index=data['time_index'])}
    #
    # node_names = {'Mixed street': ['Building' + str(i) for i in range(data['n_buildings'])],
    #               'Old street': ['Building' + str(i) for i in range(data['n_buildings'])],
    #               'New street': ['Building' + str(i) for i in range(data['n_buildings'])],
    #               'Series district': ['Street' + str(i) for i in range(data['n_streets'])],
    #               'Parallel district': ['Street' + str(i) for i in range(data['n_streets'])],
    #               'Genk': ['TermienWest', 'TermienOost', 'Boxbergheide', 'Winterslag', 'OudWinterslag',
    #                        'ZwartbergNW', 'ZwartbergZ', 'ZwartbergNE', 'WaterscheiGarden']}
    #
    # data['edge_names'] = {'Mixed street': ['dist_pipe' + str(i) for i in range(int(math.ceil(data['n_buildings'] / 2)))] +
    #                               ['serv_pipe' + str(i) for i in range(data['n_buildings'])],
    #                       'Old street': ['dist_pipe' + str(i) for i in range(int(math.ceil(data['n_buildings'] / 2)))] +
    #                             ['serv_pipe' + str(i) for i in range(data['n_buildings'])],
    #                       'New street': ['dist_pipe' + str(i) for i in range(int(math.ceil(data['n_buildings'] / 2)))] +
    #                             ['serv_pipe' + str(i) for i in range(data['n_buildings'])],
    #                       'Series district': ['dist_pipe' + str(i) for i in range(data['n_streets'])],
    #                       'Parallel district': ['dist_pipe' + str(i) for i in range(data['n_streets'])],
    #                       'Genk': ['dist_pipe' + str(i) for i in range(14)]}
    #
    # data['pipe_diameters'] = {'Mixed street': [50, 40, 32, 32, 25] + [20] * data['n_buildings'],
    #                           'Old street': [50, 40, 32, 32, 25] + [20] * data['n_buildings'],
    #                           'New street': [50, 40, 32, 32, 25] + [20] * data['n_buildings'],
    #                           'Series district': [80, 65, 50],
    #                           'Parallel district': [50, 50, 50],
    #                           'Genk': [800, 250, 450, 800, 250, 400, 700, 250, 700, 450, 400, 400, 300, 300]}
    #
    # data['mult'] = {'Mixed street': [1] * data['n_buildings'],
    #         'Old street': [1] * data['n_buildings'],
    #         'New street': [1] * data['n_buildings'],
    #         'Series district': [data['n_buildings']] * data['n_streets'],
    #         'Parallel district': [data['n_buildings']] * data['n_streets'],
    #         'Genk': [633, 746, 2363, 1789, 414, 567, 1571, 584, 2094]}
    #
    # data['model_cases'] = {
    #     'Buildings - ideal network':
    #         {
    #         'pipe_model': data['pipe_models']['NoPipes'],
    #         'time_step': data['time_steps']['StSt'],
    #         'building_model': data['building_models']['RC'],
    #         'heat_profile': None
    #         },
    #     'Buildings':
    #         {
    #             'pipe_model': data['pipe_models']['StSt'],
    #             'time_step': data['time_steps']['StSt'],
    #             'building_model': data['building_models']['RC'],
    #             'heat_profile': None
    #         },
    #     'Network':
    #         {
    #             'pipe_model': data['pipe_models']['Dynamic'],
    #             'time_step': data['time_steps']['Dynamic'],
    #             'building_model': data['building_models']['Fixed'],
    #             'heat_profile': 'Reference'
    #         },
    #     'Combined - LP':
    #         {
    #             'pipe_model': data['pipe_models']['Dynamic'],
    #             'time_step': data['time_steps']['Dynamic'],
    #             'building_model': data['building_models']['Fixed'],
    #             'heat_profile': 'Flexibility'
    #         },
    #     'Combined - MINLP':
    #         {
    #             'pipe_model': data['pipe_models']['Dynamic'],
    #             'time_step': data['time_steps']['Dynamic'],
    #             'building_model': data['building_models']['RC'],
    #             'heat_profile': None
    #         }
    # }
    #
    # save_obj(data, sim_name + '_data')