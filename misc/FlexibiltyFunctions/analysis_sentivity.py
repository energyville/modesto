import os
import pickle
import pandas as pd
import modesto.utils as ut
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pandas import ExcelWriter
import numpy as np
import math
import openpyxl
from ResponseFunctionGen import *


class SensitivityAnalysis:

    def __init__(self, name, source_dir, target_dir, sens_cases):
        self.name = name
        self.source_dir = source_dir
        self.target_dir = target_dir
        self.makedir(target_dir)

        self.district_cases = []
        self.model_cases = []
        self.flex_cases = []
        self.sens_cases = []
        self.sens_cases_names = sens_cases

        self.results = self.load_results()

        self.data = self.load_data()
        self.time_index = self.get_time_index()

        self.solutions = ['energy_use_difference', 'energy_use_ref', 'energy_use_flex',
                          'energy_cost_ref', 'energy_cost_flex', 'energy_cost_difference',
                          'upward_energy', 'downward_energy']

        self.pit = self.find_price_increase_time()

        self.output = None

    def get_time_index(self):

        for i in range(len(self.model_cases)):
            try:
                return self.results[self.sens_cases[0].name][self.district_cases[0].name]\
                    [self.model_cases[i].name][self.sens_cases[0].name][self.sens_cases[0].values[0]] \
                    [self.flex_cases[0].name]['heat_injection'].index
            except:
                pass

    def load_obj(self, file_loc):
        with open(file_loc, 'rb') as fp:
            return pickle.load(fp)

    def load_data(self):
        districts = self.results[self.sens_cases_names[0]].keys()
        models = self.results[self.sens_cases_names[0]][districts[0]].keys()
        for sens in self.sens_cases_names:
            sens_values = self.results[sens][districts[0]][models[0]][sens].keys()
        self.create_district_cases(districts)
        self.create_model_cases(models)
        self.create_sens_cases(self.sens_cases_names)

        for case in self.sens_cases:
            case.values = sens_values

        self.create_flex_cases()

    def create_district_cases(self, cases):
        available_cases = {'Old street': OldStreetCase,
                           'Mixed street': MixedStreetCase,
                           'New street': NewStreetCase,
                           'Series district': SeriesCase,
                           'Parallel district': ParallelCase,
                           'Genk': GenkCase}

        for case in cases:
            if case not in available_cases:
                raise Exception('{} is an invalid class type, choose among {} instead'.
                                format(case, available_cases.keys()))
            self.district_cases.append(available_cases[case]())

    def create_model_cases(self, cases):
        available_cases = {'Buildings - ideal network': BuildingsIdealCase,
                           'Buildings': BuildingCase,
                           'Network': NetworkCase,
                           'Combined - LP': CombinedLPCase}
        for case in cases:
            if case not in available_cases:
                raise Exception('{} is an invalid model cases, choose among {} instead'
                                .format(cases, available_cases.keys()))
            self.model_cases.append(available_cases[case]())

    def create_flex_cases(self):
        available_cases = {'Reference': ReferenceCase,
                           'Flexibility': FlexibilityCase}

        for case in available_cases:
            self.flex_cases.append(available_cases[case]())

    def create_sens_cases(self, cases):
        available_cases = {'Network size': NetworkSize,
                           'Pipe length': PipeLength,
                           'Pipe diameter': PipeDiameter,
                           'Heat demand': HeatDemand,
                           'Supply temperature level': SupplyTempLevel,
                           'Supply temperature reach': SupplyTempReach,
                           'Substation temperature difference': SubstationTempDiff}

        for case in cases:
            self.sens_cases.append(available_cases[case]())

    def load_results(self):
        results = {}
        for sens_case in self.sens_cases_names:
            results[sens_case] = self.load_obj(os.path.join(self.source_dir, self.name + '_' + sens_case + '.pkl'))

        return results

    def makedir(self, name):
        if not os.path.exists(name):
            os.makedirs(name)

    def make_dataframe(self, sens_case, index1, columns):
        index2 = sens_case.get_values()
        index1_lists = [[key]*len(index2) for key in index1]
        index1_list = [item for sublist in index1_lists for item in sublist]
        index2_lists = [index2*len(index1)]
        index2_list = [item for sublist in index2_lists for item in sublist]
        tuples = list(zip(*[index1_list, index2_list]))
        index = pd.MultiIndex.from_tuples(tuples, names=['Result type', 'Sensitivity values'])
        return pd.DataFrame(index=index, columns=columns)

    def save_xls(self, dict_dfs, file_name):
        xls_path = os.path.join(self.target_dir, file_name)
        if not os.path.isfile(xls_path):
            wb = openpyxl.Workbook()
            wb.save(xls_path)
        book = openpyxl.load_workbook(xls_path)

        for sheet in book.sheetnames:
            if 'Sheet' in sheet:
                book.remove(book[sheet])

        writer = ExcelWriter(xls_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for name, df in dict_dfs.items():
            df.to_excel(writer, name)

        writer.save()

    def save_output(self):
        if self.output is None:
            raise Exception('No output has been collected yet')

        self.save_xls(self.output, sim_name + '.xlsx')

    def savefig(self, figure, figname):
        figure.savefig(os.path.join(self.target_dir, self.name, figname + '.svg'))

    def find_result(self, sens_case=None, sens_val=None, dist_case=None, model_case=None, flex_case=None, var_name=None):
        if sens_case is None:
            return self.results
        elif dist_case is None:
            return self.results[sens_case]
        elif model_case is None:
            return self.results[sens_case][dist_case]
        elif sens_val is None:
            return self.results[sens_case][dist_case][model_case]
        elif flex_case is None:
            return self.results[sens_case][dist_case][model_case][sens_case][sens_val]
        elif var_name is None:
            return self.results[sens_case][dist_case][model_case][sens_case][sens_val][flex_case]
        else:
            return self.results[sens_case][dist_case][model_case][sens_case][sens_val][flex_case][var_name]

    def energy_use_kwh(self, heat_injection, time_step):  # TODO Resample price
        return sum(heat_injection) * time_step /1000/3600

    def difference(self, use1, use2):
        return use1 - use2

    def resample(self, data, new_time_step, last_point):
        resampled = data.resample(str(new_time_step) + 'S').pad()
        resampled = resampled.ix[~(resampled.index > last_point)]

        return resampled

    def energy_cost(self, heat_injection, price, time_step):
        price_profile = price.multiply(heat_injection) * time_step / 1000 / 3600
        return sum(price_profile[np.isfinite(price_profile)])

    def get_sens_case(self, name):
        for sens in self.sens_cases:
            if name == sens.name:
                return sens
        raise KeyError('{} is not a valid district name'.format(name))

    def get_district_case(self, name):
        for district in self.district_cases:
            if name == district.name:
                return district
        raise KeyError('{} is not a valid district name'.format(name))

    def get_flex_case(self, name):
        for flex in self.flex_cases:
            if name == flex.name:
                return flex
        raise KeyError('{} is not a valid district name'.format(name))

    def get_model_case(self, name):
        for model in self.model_cases:
            if name == model.name:
                return model
        raise KeyError('{} is not a valid model name'.format(name))

    def find_price_increase_time(self):
        return self.time_index[int(self.get_flex_case('Flexibility').get_pos() * len(self.time_index))]

    def find_time_step(self, model_case):
        return model_case.get_time_step()

    def find_last_point(self):
        return self.time_index[-1]

    def upward_downward_power_kwh(self, heat_injection_flex, heat_injection_ref, pit, time_step):
        response = self.difference(heat_injection_flex,
                                   heat_injection_ref)

        return sum(response.ix[response.index < pit]) * time_step / 1000 / 3600, \
               -sum(response.ix[response.index >= pit]) * time_step / 1000 / 3600

    def get_price(self, flex_case_name):
        return self.get_flex_case(flex_case_name).get_price_profile(self.time_index)

    def collect_results(self, sens_case_names=None, dist_case_names=None, model_case_names=None):
        if sens_case_names is None:
            sens_cases = self.sens_cases
        else:
            sens_cases = []
            for name in sens_case_names:
                sens_cases.append(self.get_sens_case(name))

        if dist_case_names is None:
            dist_cases = self.district_cases
        else:
            dist_cases = []
            for name in dist_case_names:
                dist_cases.append(self.get_district_case(name))

        if model_case_names is None:
            model_cases = self.model_cases
        else:
            model_cases = []
            for name in model_case_names:
                model_cases.append(self.get_model_case(name))

        output = {}

        for sens in sens_cases:
            output[sens.name] = self.make_dataframe(sens, self.solutions, [m.name for m in model_cases])

        for sens in sens_cases:
            for sens_val in sens.get_values():
                for dist in dist_cases:
                    for model in self.model_cases:

                        def change_element(df, new_val, sol_type):
                            df.loc[(sol_type, sens_val), model.name] = new_val
                            return df

                        def get_element(df, sol_type):
                            return df.loc[(sol_type, sens_val), model.name]

                        case_results_flex = self.find_result(sens_case=sens.name,
                                                             dist_case=dist.name,
                                                             model_case=model.name,
                                                             flex_case='Flexibility',
                                                             sens_val=sens_val)
                        case_results_ref = self.find_result(sens_case=sens.name,
                                                            dist_case=dist.name,
                                                            model_case=model.name,
                                                            flex_case='Reference',
                                                            sens_val=sens_val)

                        if case_results_flex is not None and case_results_ref is not None:
                            sens_output = output[sens.name]
                            time_step = self.find_time_step(model)
                            price = self.resample(self.get_price('Flexibility'), time_step, self.find_last_point())
                            sens_output = change_element(
                                sens_output,
                                self.energy_use_kwh(case_results_ref['heat_injection'], time_step), 'energy_use_ref')
                            sens_output = change_element(
                                sens_output,
                                self.energy_use_kwh(case_results_flex['heat_injection'], time_step), 'energy_use_flex')
                            sens_output = change_element(
                                sens_output,
                                self.difference(get_element(sens_output, 'energy_use_flex'),
                                                get_element(sens_output, 'energy_use_ref')),
                                'energy_use_difference')
                            sens_output = change_element(
                                sens_output,
                                self.energy_cost(case_results_ref['heat_injection'], price, time_step),
                                'energy_cost_ref')
                            sens_output = change_element(
                                sens_output,
                                self.energy_cost(case_results_flex['heat_injection'], price, time_step),
                                'energy_cost_flex')
                            sens_output = change_element(
                                sens_output,
                                self.difference(get_element(sens_output, 'energy_cost_ref'),
                                                get_element(sens_output, 'energy_cost_flex')),
                                'energy_cost_difference')

                            upward_energy, downward_energy = self.upward_downward_power_kwh(
                                case_results_flex['heat_injection'], case_results_ref['heat_injection'], self.pit, time_step)
                            sens_output = change_element(sens_output, upward_energy,'upward_energy')
                            sens_output = change_element(sens_output, downward_energy, 'downward_energy')

                        else:
                            print 'WARNING: There is no data for case {}: {} - {} - {}'.format(sens.name, sens_val,
                                                                                               dist.name, model.name)
        self.output = output
        return output


if __name__ == '__main__':

    sim_name = 'sensitivity_280918'
    source_dir = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'results')
    target_dir = os.path.join(source_dir, sim_name)

    analysis = SensitivityAnalysis(sim_name, source_dir, target_dir, ['Supply temperature level'])
    print analysis.collect_results()
    analysis.save_output()

    print analysis.results['Supply temperature level']['Genk'] \
        ['Buildings']['Supply temperature level'][323.15]['Flexibility']['day_zone_temperatures']
    plt.plot(analysis.results['Supply temperature level']['Genk'] \
        ['Buildings']['Supply temperature level'][323.15]['Flexibility']['day_zone_temperatures'])
    plt.show()
